# -*- coding: utf-8 -*-
"""
Created on Thu Aug  6 15:51:29 2026

@author: ricky
"""

import pandas as pd
from sqlalchemy import create_engine
from datetime import datetime
from openpyxl.styles import Font


# --- Parametri di connessione MySQL: modifica con i tuoi valori ---
DB_USER = "root"
DB_PASSWORD = "Lunabred.01"
DB_HOST = "localhost"
DB_PORT = "3306"
DB_NAME = "supply_chain"

ENGINE_URL = f"mysql+mysqlconnector://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"


queries_kpi = {
       "ritardo per regione": 
           """SELECT order_region,
               ROUND(AVG(days_for_shipping_real), 2) AS giorni_reali_medi,
               ROUND(AVG(days_for_shipment_scheduled), 2) AS giorni_schedulati_medi,
               ROUND(AVG(days_for_shipping_real - days_for_shipment_scheduled), 2) AS ritardo_medio
               FROM orders
               GROUP BY order_region
               ORDER BY ritardo_medio DESC
               LIMIT 10""",

       "rischio ritardo per modalita di spedizione":
           """SELECT shipping_mode,
               COUNT(*) AS totale_ordini,
               SUM(CASE WHEN late_delivery_risk = 1 THEN 1 ELSE 0 END) AS ordini_a_rischio_ritardo,
               ROUND(100.0 * SUM(CASE WHEN late_delivery_risk = 1 THEN 1 ELSE 0 END) / COUNT(*), 1) AS perc_a_rischio
               FROM orders
               GROUP BY shipping_mode
               ORDER BY perc_a_rischio DESC""",

        "Top 10 categorie prodotto per fatturato":
            """SELECT category_name,
                ROUND(SUM(sales), 2) AS fatturato_totale,
                COUNT(*) AS numero_ordini
                FROM orders
                GROUP BY category_name
                ORDER BY fatturato_totale DESC
                LIMIT 10""",

        "Fatturato mensile":
            """SELECT DATE_FORMAT(STR_TO_DATE(order_date_dateorders, '%c/%e/%Y %H:%i'), '%Y-%m') AS mese,
                ROUND(SUM(sales), 2) AS fatturato_totale
                FROM orders
                GROUP BY DATE_FORMAT(STR_TO_DATE(order_date_dateorders, '%c/%e/%Y %H:%i'), '%Y-%m')
                ORDER BY mese""",

        "Top 10 clienti per numero di ordini":
            """SELECT customer_id, COUNT(*) AS numero_ordini
                FROM orders
                GROUP BY customer_id
                ORDER BY numero_ordini DESC
                LIMIT 10"""
}
    
def get_engine():
    return create_engine(ENGINE_URL)

def get_kpi_data(engine, query):
    df = pd.read_sql_query(query, engine)
    return df

if __name__ == "__main__":
    engine = get_engine()
    oggi = datetime.today().strftime("%Y-%m-%d")
    nome_file = f"report_kpi_{oggi}.xlsx"

with pd.ExcelWriter(nome_file) as writer:
        for nome_foglio, query in queries_kpi.items():
            df = get_kpi_data(engine, query)
            df.to_excel(writer, sheet_name=nome_foglio, index=False)

from openpyxl import load_workbook

wb = load_workbook(nome_file)

for foglio in wb.sheetnames:
    ws = wb[foglio]
    # Grassetto sulla prima riga (intestazioni)
    for cella in ws[1]:
        cella.font = Font(bold=True)
    # Allarga le colonne in base al contenuto piu' lungo
    for colonna in ws.columns:
        lunghezza_max = max(len(str(cella.value)) for cella in colonna)
        lettera_colonna = colonna[0].column_letter
        ws.column_dimensions[lettera_colonna].width = lunghezza_max + 2

wb.save(nome_file)

print(f"Report creato: {nome_file}")        
